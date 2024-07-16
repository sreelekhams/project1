from django.shortcuts import render, redirect, get_object_or_404
from .models import Department,User,Designation,Location,Employee,Skill
from .forms import DepartmentForm,Designation_Add_Form,LocationForm,EmployeeForm,SkillFormSet,User_Add_Form,User_Edit_Form,UploadFileForm
from django.contrib.auth import authenticate, login,logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import JsonResponse
from django.conf import settings
from django.contrib.auth.hashers import make_password, check_password
from django.http import HttpResponse
from datetime import datetime
import openpyxl
import os
from master.forms import *
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from master.databsequery import *
import logging
from django.core.mail import EmailMessage
import threading
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from io import BytesIO
from django.core.mail import EmailMessage, BadHeaderError
import logging
logger = logging.getLogger(__name__)
from django.contrib import messages
import uuid
import json




# Create your views here.
def indexpage(request):
    return render(request, 'index.html')



def user_login(request):

    template_name = 'login.html'

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_exist = User.objects.filter(username=username).exists()
       
        if user_exist:
           
            user = authenticate(request, username=username, password=password)
           
            if user is not None:
                if user.role == 'ADMIN' :
                   
                    login(request, user)
           
                    return redirect('indexpage')
                
                elif user.role == 'VIEWER':
                    login(request, user)
                    return redirect('indexpage')
                
                else:
                    context = {'msg': 'Invalid Username or Password!'}
                    return render(request, template_name, context)
            else:
                
                context = {'msg': 'Password is incorrect!'}
                return render(request, template_name, context)

        else:
            context = {'msg': 'User Does Not exist'}
            return render(request, template_name, context)  
            
    return render(request, template_name)


def admin_logout(request):
    
    logout(request)
  
    return redirect(user_login)

@login_required(login_url='adlogin')
def department_add(request):
    form = DepartmentForm()
    template_name = 'master/add_department.html'
    context = {'form': form}
    
    if request.method == 'POST':
        print(request.user.id,"Form submitted")
        form = DepartmentForm(request.POST, request.FILES)
        
        if form.is_valid():
            print("Form is valid")
            data = form.save()
            data.created_by =User.objects.get(id=request.user.id)
            data.save()
           
            messages.success(request, 'Department Successfully added.', 'alert-success')
            return redirect('department_list')
            
        else:
            print("Form is not valid")
            print(form.errors)  # Print form errors to debug
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else :
        print("Rendering form")
        return render(request, template_name, context)
    
@login_required(login_url='adlogin')    
def department_edit(request, pk):
    template_name = 'master/department_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Department not found.', 'alert-danger')
        return redirect('department_list')
    try:
        dep_obj = Department.objects.get(department_id=pk)
    except Department.DoesNotExist:
        messages.error(request, 'Department not found.', 'alert-danger')
        return redirect('department_list')
    form = DepartmentForm(instance=dep_obj)
    context = {'form': form, 'dep_obj': dep_obj}
    if request.method == 'POST':
        form = DepartmentForm(request.POST, request.FILES, instance=dep_obj)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            messages.success(request, 'Department Successfully Updated.', 'alert-success')
            return redirect('department_list')
        else:
            print(form.errors)
            messages.error(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    
# @login_required(login_url='adlogin')
# def department_list(request):
#     # Write your raw SQL query
#     sql_query = "SELECT * FROM master_department;"
    
#     # Execute the query
#     with connection.cursor() as cursor:
#         cursor.execute(sql_query)
#         departments = cursor.fetchall()  # Fetch all rows
   
#     # Process the results
#     context = {
#         'departments': departments
#     }
    
#     return render(request, 'master/department_list.html', context)


@login_required(login_url='adlogin')
def department_detail(request,pk):
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Invalid department ID.', 'alert-danger')
        return redirect('department_list')

    try:
        departments = Department.objects.get(department_id=pk)
    except departments.DoesNotExist:
        messages.error(request, 'Department not found.', 'alert-danger')
        return redirect('department_list')
    context = {
        'departments': departments
    }
    
    return render(request, 'master/department_detail.html', context)

@login_required(login_url='adlogin')
def department_delete(request, pk):
    department = Department.objects.get(department_id=pk)
    
    department.delete()
    messages.success(request, 'Department Deleted Successfully', 'alert-success')
    return redirect('department_list')

@login_required(login_url='adlogin')
def designation_add(request):
    form = Designation_Add_Form
    template_name = 'master/add_designation.html'
    context = {'form': form}
   
    if request.method == 'POST':
        print(request.user.id,"Form submitted")
        form = Designation_Add_Form(request.POST, request.FILES)
        
        if form.is_valid():
            print("Form is valid")
            data = form.save()
            data.created_by =User.objects.get(id=request.user.id)
            data.save()
           
            messages.success(request, 'Designation Successfully Updated.', 'alert-success')
            return redirect('designation_list')
            
        else:
            print("Form is not valid")
            print(form.errors)  # Print form errors to debug
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else :
        print("Rendering form")
        return render(request, template_name, context)
    

@login_required(login_url='adlogin')   
def designation_edit(request, pk):
    template_name = 'master/designation_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Designation not found.', 'alert-danger')
        return redirect('designation_list')
    try:
        des_obj = Designation.objects.get(designation_id=pk)
    except des_obj.DoesNotExist:
        messages.error(request, 'Designation not found.', 'alert-danger')
        return redirect('designation_list')
    form = Designation_Add_Form(instance=des_obj)
    context = {'form': form, 'des_obj': des_obj}
    if request.method == 'POST':
        form = Designation_Add_Form(request.POST, request.FILES, instance=des_obj)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            messages.success(request, 'Designation Successfully Updated.', 'alert-success')
            return redirect('designation_list')
        else:
            print(form.errors)
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
@login_required(login_url='adlogin')    
# def designation_list(request):
   
#     sql_query = """
#         SELECT d.designation_id, d.designation_name, d.description, dep.department_name
#         FROM master_designation d
#         LEFT JOIN master_department dep ON d.department_id = dep.department_id;
#     """
    
   
#     with connection.cursor() as cursor:
#         cursor.execute(sql_query)
#         designations = cursor.fetchall()  # Fetch all rows

   
#     designation_list = [
#         {
#             'designation_id': row[0],
#             'designation_name': row[1],
#             'description': row[2],
#             'department_name': row[3]
#         }
#         for row in designations
#     ]
#     print(designation_list,"designation_list")
   
#     context = {
#         'designation_list': designation_list
#     }
    
#     return render(request, 'master/designation_list.html', context)

@login_required(login_url='adlogin')
def designation_detail(request,pk):
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'designation not found.', 'alert-danger')
        return redirect('designation_list')

    try:
         designation = get_object_or_404(Designation,designation_id=pk)
    except designation.DoesNotExist:
        messages.error(request, 'designation not found.', 'alert-danger')
        return redirect('designation_list')
    
    department_name = designation.department.department_name
    context = {
        'designation': designation,
        'department_name':department_name
    }
    
    return render(request, 'master/designation_detail.html', context)

@login_required(login_url='adlogin')
def designation_delete(request, pk):
    designation = Designation.objects.get(designation_id=pk)
    
    designation.delete()
    messages.success(request, 'Designation Deleted Successfully', 'alert-success')
    return redirect('designation_list')

@login_required(login_url='adlogin')
def location_add(request):
    form = LocationForm
    template_name = 'master/add_location.html'
    context = {'form': form}
   
    if request.method == 'POST':
        print(request.user.id,"Form submitted")
        form = LocationForm(request.POST, request.FILES)
        
        if form.is_valid():
            print("Form is valid")
            data = form.save()
            data.created_by =User.objects.get(id=request.user.id)
            data.save()
            messages.success(request, 'Location Added Successfully', 'alert-success')
          
            return redirect('location_list')
            
        else:
            print("Form is not valid")
            print(form.errors)  # Print form errors to debug
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else :
        print("Rendering form")
        return render(request, template_name, context)
    
# @login_required(login_url='adlogin')
# def location_list(request):
   
#     sql_query = "SELECT * FROM master_location;"
    
   
#     with connection.cursor() as cursor:
#         cursor.execute(sql_query)
#         location = cursor.fetchall()  
#     print(location,"departments")
   
#     context = {
#         'location': location
#     }
    
#     return render(request, 'master/location_list.html', context)


@login_required(login_url='adlogin')
def location_detail(request,pk):
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')

    try:
         location = get_object_or_404(Location,location_id=pk)
    except location.DoesNotExist:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')
    
    
    
    context = {
        'location': location
    }
    
    return render(request, 'master/location_detail.html', context)

@login_required(login_url='adlogin')
def location_edit(request, pk):
    template_name = 'master/location_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')
    try:
        loc_obj = Location.objects.get(location_id=pk)
    except loc_obj.DoesNotExist:
        messages.error(request, 'Location not found.', 'alert-danger')
        return redirect('location_list')
    
    form = LocationForm(instance=loc_obj)
    context = {'form': form, 'loc_obj': loc_obj}
    if request.method == 'POST':
        form = LocationForm(request.POST, request.FILES, instance=loc_obj)
        if form.is_valid():
            data = form.save(commit=False)
            data.save()
            messages.success(request, 'Location Successfully Updated.', 'alert-success')
            return redirect('location_list')
        else:
            print(form.errors)
            messages.success(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    

@login_required(login_url='adlogin')    
def location_delete(request, pk):
    location = Location.objects.get(location_id=pk)
    
    location.delete()
    messages.success(request, 'Location Deleted Successfully', 'alert-success')
    return redirect('location_list')

@login_required(login_url='adlogin')
def employee_add(request):
    form = EmployeeForm
    formset = SkillFormSet(queryset=Skill.objects.none())
    template_name = 'master/add_employee.html'
    context = {'form': form, 'formset': formset}
   
    if request.method == 'POST':
        print(request.user.id,"Form submitted")
        form = EmployeeForm(request.POST, request.FILES)
        formset = SkillFormSet(request.POST, queryset=Skill.objects.none())
        if form.is_valid() and formset.is_valid():
            print("Form is valid")
            data = form.save()
            data.created_by =User.objects.get(id=request.user.id)
            data.save()
            
            
            for skill_form in formset:
                skill = skill_form.save(commit=False)
                skill.employee = data
                skill.save()
            messages.success(request, 'Employee Added Successfully', 'alert-success')
            return redirect('employee_list')
            
        else:
            print("Form is not valid")
            print(form.errors)  
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form,'formset': formset}
            return render(request, template_name, context)
    else :
        print("Rendering form")
        return render(request, template_name, context)
    
def designations(request):
    department_id = request.GET.get('department')
    designations = Designation.objects.filter(department=department_id).all()
    return JsonResponse(list(designations.values('designation_id', 'designation_name')), safe=False)

# def employee_list(request):
#     # Write your raw SQL query
#     sql_query = """
#     SELECT 
#         e.employee_id,e.join_date, e.emp_no, e.name, e.phone, e.address, 
#         e.emp_start_date, e.emp_end_date,e.photo,e.status,
#         d.department_name, ds.designation_name, l.location_name
#     FROM master_employee e
#     LEFT JOIN master_department d ON e.department_id = d.department_id
#     LEFT JOIN master_designation ds ON e.designation_id = ds.designation_id
#     LEFT JOIN master_location l ON e.location_id = l.location_id;
#     """
    
   
#     with connection.cursor() as cursor:
#         cursor.execute(sql_query)
#         employees = cursor.fetchall()  # Fetch all rows
    
   
#     employee_list = []
#     for row in employees:
#         employee = {
#             'employee_id':row[0],
#             'join_date': row[1],
#             'emp_no': row[2],
#             'name': row[3],
#             'phone': row[4],
#             'address': row[5],
#             'emp_start_date': row[6],
#             'emp_end_date': row[7],
#             'photo':settings.MEDIA_URL + row[8],
#             'status':row[9],
#             'department_name': row[10],
#             'designation_name': row[11],
#             'location_name': row[12],
#         }
#         employee_list.append(employee)
   
#     context = {
#         'employees': employee_list
#     }
    
#     return render(request, 'master/employee_list.html', context)

@login_required(login_url='adlogin')
def employee_edit(request, pk):
    template_name = 'master/employee_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'Employee not found.', 'alert-danger')
        return redirect('employee_list')
    try:
        emp_obj = get_object_or_404(Employee, employee_id=pk)
    except emp_obj.DoesNotExist:
        messages.error(request, 'Employee not found.', 'alert-danger')
        return redirect('employee_list')
   
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=emp_obj)
        formset = SkillFormSet(request.POST, queryset=Skill.objects.filter(employee=emp_obj))
        
        if form.is_valid() :
         
            employee = form.save(commit=False)
            employee.save()
           
            if formset.is_valid():
                for i in formset:
                  
                    skill = i.save()
                    skill.employee = employee
                    skill.save()
             
            messages.success(request, 'Employee Successfully Updated.', 'alert-success')
            return redirect('employee_list')
        else:
          
            messages.error(request, 'Data is not valid.', 'alert-danger')
    else:
        form = EmployeeForm(instance=emp_obj)
        formset = SkillFormSet(queryset=Skill.objects.filter(employee=emp_obj))
    
    context = {'form': form, 'formset': formset, 'emp_obj': emp_obj}
    return render(request, template_name, context)



@login_required(login_url='adlogin')
def employee_detail(request,pk):
     
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'employee not found.', 'alert-danger')
        return redirect('employee_list')

    try:
         employee = get_object_or_404(Employee,employee_id=pk)
    except employee.DoesNotExist:
        messages.error(request, 'employee not found.', 'alert-danger')
        return redirect('employee_list')
    
    department=employee.department.department_name
    designation=employee.designation.designation_name
    location=employee.location.location_name
    skills = Skill.objects.filter(employee=employee)
    
    context = {
        
        'employee': employee,
        'department':department,
        'location':location,
        'designation':designation,
        'skills': skills,
    }
    
    return render(request, 'master/employee_detail.html', context)
   
@login_required(login_url='adlogin')    
def employee_delete(request, pk):
    employee = Employee.objects.get(employee_id=pk)
    
    employee.delete()
    messages.success(request, 'Employee Deleted Successfully', 'alert-success')
    return redirect('employee_list')


def user_add(request):
    form = User_Add_Form
   
    template_name = 'accounts/user_add.html'
    
    context = {'form': form}
    if request.method == 'POST':
        form = User_Add_Form(request.POST, request.FILES)
        
        if form.is_valid() :
            data = form.save(commit=False)
           
            passw = data.password
            passw = make_password(passw)
            data.password = passw
           
            data.save()
            messages.success(request, 'User Added Successfully', 'alert-success')
            return redirect('user_list')
        else:
            messages.error(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form,}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    

@login_required(login_url='adlogin')
@csrf_exempt
def user_list(request):
    if request.method == "GET":
        template_name = 'accounts/user_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
       
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        des = user_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(des)


@login_required(login_url='adlogin')
def user_detail(request,pk):
    try:
       
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'user not found.', 'alert-danger')
        return redirect('user_list')

    try:
         user = get_object_or_404(User,id=pk)
    except user.DoesNotExist:
        messages.error(request, 'user not found.', 'alert-danger')
        return redirect('user_list')
     
    
    context = {
        'user': user
    }
    
    return render(request, 'accounts/user_detail.html', context)


def user_edit(request, pk):
    template_name = 'accounts/user_edit.html'
    try:
        uuid_obj = uuid.UUID(pk)
    except ValueError:
        messages.error(request, 'user not found.', 'alert-danger')
        return redirect('user_list')
    try:
        user_obj = User.objects.get(id=pk)
    except user_obj.DoesNotExist:
        messages.error(request, 'user not found.', 'alert-danger')
        return redirect('user_list')
    
    
   
    form = User_Edit_Form(instance=user_obj)
   
    context = {'form': form}
    if request.method == 'POST':
        form = User_Edit_Form(request.POST, request.FILES, instance=user_obj)
       
        if form.is_valid():
            data = form.save(commit=False)
           
            data.save()
            messages.success(request, 'User Updated Successfully', 'alert-success')
            return redirect('user_list')
        else:
            
            context = {'form': form,}
            print(form.errors)
            messages.error(request, 'Data is not valid.', 'alert-danger')
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    


@login_required(login_url='adlogin')    
def user_delete(request, pk):
    user = User.objects.get(id=pk)
    
    user.delete()
    messages.success(request, 'User Deleted Successfully', 'alert-success')
    return redirect('user_list')


def employee_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    employees = Employee.objects.all()
    

    employees_skills = []
    for employee in employees:
        skills = employee.skills.all()
        employees_skills.append({
            'employee': employee,
            'skills': skills
        })

    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            employees = employees.filter(join_date__range=[start_date, end_date])
        except ValueError:
            # Handle invalid date format
            pass

    if 'download' in request.GET:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=employee_report.xlsx'},
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Employee Report'

        # Write the header
        headers = ['Employee Number', 'Name', 'Phone', 'Start Date', 'End Date', 'Status', 'Department', 'Designation', 'Location', 'Skills']
        sheet.append(headers)

        # Write the data
        for employee in employees:
          
            skills = Skill.objects.filter(employee=employee)
          
            skills_list = ", ".join([skill.skill_name for skill in skills])
            
            row = [
                employee.emp_no,
                employee.name,
                employee.phone,
                employee.emp_start_date.strftime('%Y-%m-%d') if employee.emp_start_date else '',
                employee.emp_end_date.strftime('%Y-%m-%d') if employee.emp_end_date else '',
                employee.status,
                employee.department.department_name if employee.department else '',
                employee.designation.designation_name if employee.designation else '',
                employee.location.location_name if employee.location else '',
                skills_list,
            ]
            sheet.append(row)
       

        workbook.save(response)
        return response
  
    context = {
        'employees': employees_skills,
        
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'master/employee_report.html', context)



def handle_uploaded_file(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data

def bulk_upload_dep(request):
   
    if request.method == 'POST':
       
        data = handle_uploaded_file(request.FILES['filename'])
      
        for row in data:
            
            if not Department.objects.filter(department_name=row[0]).exists():
                Department.objects.create(department_name=row[0],description=row[1])
        return redirect('department_list')
    
   
def bulk_upload_des(request):
    if request.method == 'POST':
       
        data = handle_uploaded_file(request.FILES['file'])
        print(data,"data")
       
        for row in data:
            department_name = row[0]
            designation_name = row[1]
            description = row[2]

         
            department = Department.objects.filter(department_name=department_name).first()
            if department and not Designation.objects.filter(designation_name=designation_name, department=department).exists():
                Designation.objects.create(department=department, designation_name=designation_name, description=description)

        return redirect('designation_list')
   
def bulk_upload_loc(request):
   
    if request.method == 'POST':
       
        data = handle_uploaded_file(request.FILES['file'])
      
        for row in data:
            
            if not Location.objects.filter(location_name=row[0]).exists():
                Location.objects.create(location_name=row[0],description=row[1])
        return redirect('location_list')
    

def bulk_upload_user(request):
   
    if request.method == 'POST':
       
        data = handle_uploaded_file(request.FILES['file'])
      
        for row in data:
            username=row[0]
            email=row[1] 
            first_name=row[2]
            last_name=row[3] 
            password=row[4] 
            role =row[5]
            
            if not User.objects.filter(username=username).exists():
                User.objects.create( username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    role=role,
                    password= make_password(password))
                
            else:
                # Log the username that already exists
                logging.warning(f"Username {username} already exists. Skipping this entry.")
        return redirect('user_list')
    

def save_uploaded_file(uploaded_file, save_directory):
    try:
      
        file_name, uploaded_file = next(iter(uploaded_file.items()))
        if not os.path.exists(save_directory):
            os.makedirs(save_directory)

        # Construct the full path to save the file
        save_path = os.path.join(save_directory, uploaded_file.name)
        print(save_path,'save')
        # Save the file to the specified directory
        with default_storage.open(save_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        print(f"File saved at {save_path}")

    except Exception as e:
        print(f"An error occurred: {e}")


def bulk_upload_emp(request):
    if request.method == 'POST':
        files = request.FILES.getlist('files')  # Accessing multiple files using getlist()
        print(files)
        photo_files = {file.name: file for file in files if file.name.endswith(('.png', '.jpg', '.jpeg'))} 
        save_directory = 'employee_photos'

        save_uploaded_file(photo_files, save_directory)
        print(photo_files,"photo_files") # Collect photo files
        errors = []
        for uploaded_file in files:
            try:
                data = handle_uploaded_file(uploaded_file)  # Process each file using your utility function

                for row in data:
                    join_date = row[0]
                    emp_no = row[1]
                    name = row[2]
                    phone = row[3]
                    address = row[4]
                    emp_start_date = row[5]
                    emp_end_date = row[6]
                    status = row[7]
                    department_name = row[8]
                    designation_name = row[9]
                    location_name = row[10]
                    photo_name  = row[11]
                    skills_data = row[12:]  # Assuming skills data starts from the 12th column onwards

                  
                    employee_exists = Employee.objects.filter(emp_no=emp_no).exists()
                    if employee_exists:
                        raise ValueError(f'Employee with Employee No. {emp_no} already exists.')

                    department = Department.objects.filter(department_name=department_name).first()
                    designation = Designation.objects.filter(designation_name=designation_name, department=department).first()
                    location = Location.objects.filter(location_name=location_name).first()

                    if not department or not designation or not location:
                        raise ValueError('Invalid department, designation, or location')

                   
                  
                    photo_file = photo_files.get(photo_name)
                    print(photo_file)
                   
                    employee = Employee.objects.create(
                        emp_no=emp_no,
                        join_date=join_date,
                        name=name,
                        phone=phone,
                        address=address,
                        emp_start_date=emp_start_date,
                        emp_end_date=emp_end_date,
                        status=status,
                        department=department,
                        designation=designation,
                        location=location,
                        photo='employee_photos'+'/'+photo_name
                    )

                    # Create associated skills
                    for i in range(0, len(skills_data), 2):  # Assuming each skill has two columns: skill_name and description
                        skill_name = skills_data[i]
                        skill_description = skills_data[i + 1]
                        Skill.objects.create(employee=employee, skill_name=skill_name, description=skill_description)
                

            except Exception as e:
                errors.append(str(e))

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            messages.success(request, 'Employees uploaded successfully.')

        return redirect('employee_list')

    return render(request, 'employee_list.html')


def export_departments_to_excel(request):
   
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Departments'


    columns = ['Sl.No',  'Department Name', 'Description']
    row_num = 1

  
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    
    for index, department in enumerate(Department.objects.all(), start=1):
        row_num += 1
        row = [
            index,  # Sl.No
            department.department_name,
            department.description,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departments.xlsx'

    workbook.save(response)
    return response



def export_designations_to_excel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Designations'

    columns = ['Sl.No', 'Department Name', 'Designation Name', 'Description']
    row_num = 1

   
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   
    for index, designation in enumerate(Designation.objects.all(), start=1):
        row_num += 1
        row = [
            index, 
            designation.department.department_name if designation.department else '',
            designation.designation_name,
            designation.description,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=designations.xlsx'

    workbook.save(response)
    return response


def export_locations_to_excel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Locations'

    columns = ['Sl.No', 'Location Name', 'Description']
    row_num = 1

   
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   
    for index, location in enumerate(Location.objects.all(), start=1):
        row_num += 1
        row = [
            index,  
            location.location_name,
            location.description,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=locations.xlsx'

    workbook.save(response)
    return response


def export_employees_to_excel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Employees'

  
    columns = ['Sl.No', 'Employee No', 'Join Date', 'Name', 'Phone', 'Address', 
               'Emp Start Date', 'Emp End Date', 'Status', 'Department', 'Designation', 'Location', 'Skills']

  
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

   
    employees = Employee.objects.all()

   
    for index, employee in enumerate(employees, start=2):
        skills = ', '.join([skill.skill_name for skill in employee.skills.all()])
        row = [
            index - 1, 
            employee.emp_no,
            employee.join_date.strftime('%Y-%m-%d'),
            employee.name,
            employee.phone,
            employee.address,
            employee.emp_start_date.strftime('%Y-%m-%d'),
            employee.emp_end_date.strftime('%Y-%m-%d'),
            employee.status,
            employee.department.department_name if employee.department else '',  
            employee.designation.designation_name if employee.designation else '',  
            employee.location.location_name if employee.location else '', 
            skills, 
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=index, column=col_num)
            cell.value = cell_value

   
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'

 
    workbook.save(response)

    return response


def export_users(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Users'

   
    columns = ['Sl.No', 'Username', 'Email', 'First Name', 'Last Name']
    row_num = 1

  
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

   
    for index, user in enumerate(User.objects.all(), start=1):
        row_num += 1
        row = [
            index,
            user.username,
            user.email,
            user.first_name,
            user.last_name,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

   
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'

    workbook.save(response)
    return response


@login_required(login_url='adlogin')
@csrf_exempt
def employee_list(request):
    if request.method == "GET":
        template_name = 'master/employee_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        emp = emp_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(emp)
    





@login_required(login_url='adlogin')
@csrf_exempt
def department_list(request):
    if request.method == "GET":
        template_name = 'master/department_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        dep = department_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(dep)
    

    



@login_required(login_url='adlogin')
@csrf_exempt
def location_list(request):
    if request.method == "GET":
        template_name = 'master/location_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        loc = location_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(loc)

@login_required(login_url='adlogin')
@csrf_exempt
def designation_list(request):
    if request.method == "GET":
        template_name = 'master/designation_list.html'
       
        return render(request, template_name, )

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        des = designation_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(des)
    
def generate_employee_pdf(employee_id):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    employee = Employee.objects.get(employee_id=employee_id)

    # Add a heading with a larger font
    p.setFont("Helvetica-Bold", 18)
    p.drawString(100, height - 50, "Employee Details")

 
    p.setFont("Helvetica-Bold", 12)
    p.drawString(70, height - 90, "Employee Name:")
    p.drawString(70, height - 110, "Employee Number:")
    p.drawString(70, height - 130, "Join Date:")
    p.drawString(70, height - 150, "Phone:")
    p.drawString(70, height - 170, "Address:")
    p.drawString(70, height - 190, "Start Date:")
    p.drawString(70, height - 210, "End Date:")
    p.drawString(70, height - 230, "Status:")
    p.drawString(70, height - 250, "Department:")
    p.drawString(70, height - 270, "Designation:")
    p.drawString(70, height - 290, "Location:")

    # Add employee data with regular font
    p.setFont("Helvetica", 12)
    p.drawString(200, height - 90, str(employee.name))
    p.drawString(200, height - 110, str(employee.emp_no))
    p.drawString(200, height - 130, str(employee.join_date))
    p.drawString(200, height - 150, str(employee.phone))
    p.drawString(200, height - 170, str(employee.address))
    p.drawString(200, height - 190, str(employee.emp_start_date))
    p.drawString(200, height - 210, str(employee.emp_end_date) if employee.emp_end_date else 'N/A')
    p.drawString(200, height - 230, str(employee.status))
    p.drawString(200, height - 250, str(employee.department))
    p.drawString(200, height - 270, str(employee.designation))
    p.drawString(200, height - 290, str(employee.location))

    # Add employee photo with a label
    if employee.photo:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(70, height - 320, "Employee Photo:")
        p.drawImage(employee.photo.path, 200, height - 440, width=100, preserveAspectRatio=True)
       

    # Footer
    p.setFont("Helvetica-Oblique", 10)
    p.drawString(70, 50, "Generated by datahub technologies")

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer.getvalue()



def download_employee_pdf(request, employee_id):
    employee = get_object_or_404(Employee, employee_id=employee_id)
    buffer = generate_employee_pdf(employee_id)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{employee.name}_details.pdf"'
    return response



def send_pdf(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
       
        recipient_email = request.POST.get('email')
      
        employee = Employee.objects.get(employee_id=employee_id)
       
        pdf = generate_employee_pdf(employee_id)
       

        email = EmailMessage(
            subject=f'Employee Details for {employee.name}',
            body='Please find attached the employee details.',
            to=[recipient_email]
        )
       
        email.attach(f'{employee.name}_details.pdf', pdf, 'application/pdf')
       
        try:
            email.send()
           
            logger.info(f"Email sent successfully to {recipient_email}")
            return JsonResponse({'status': 'success'})
        except BadHeaderError:
            logger.error("Invalid header found.")
            return JsonResponse({'status': 'error', 'message': 'Invalid header found.'}, status=400)
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)



    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)



def record_audio(request):
    if request.method == 'POST':
        form = AudioForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Audio recorded Successfully.', 'alert-success')
            return redirect('indexpage')
    else:
        form = AudioForm()
    return render(request, 'master/record_audio.html', {'form': form})

def capture_image(request):
    if request.method == 'POST':
        form = ImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Image Captured Successfully.', 'alert-success')
            return redirect('indexpage')
    else:
        form = ImageForm()
    return render(request, 'master/capture_image.html', {'form': form})


def download_selected(request):
    if request.method == 'POST':
        try:
            employee_ids_json = request.POST.get('employee_ids')
            employee_ids = json.loads(employee_ids_json)
            valid_employee_ids = []

            for emp_id in employee_ids:
                try:
                    valid_employee_ids.append(uuid.UUID(emp_id))
                except ValueError:
                    return JsonResponse({'error': f'Invalid UUID: {emp_id}'}, status=400)
            
            employees = Employee.objects.filter(employee_id__in=valid_employee_ids)
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Employees'

            columns = ['Sl.No', 'Employee No', 'Join Date', 'Name', 'Phone', 'Address', 
                       'Emp Start Date', 'Emp End Date', 'Status', 'Department', 'Designation', 'Location', 'Skills']

            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for index, employee in enumerate(employees, start=2):
                skills = ', '.join([skill.skill_name for skill in employee.skills.all()])
                row = [
                    index - 1, 
                    employee.emp_no,
                    employee.join_date.strftime('%Y-%m-%d') if employee.join_date else '',
                    employee.name,
                    employee.phone,
                    employee.address,
                    employee.emp_start_date.strftime('%Y-%m-%d') if employee.emp_start_date else '',
                    employee.emp_end_date.strftime('%Y-%m-%d') if employee.emp_end_date else '',
                    employee.status,
                    employee.department.department_name if employee.department else '',  
                    employee.designation.designation_name if employee.designation else '',  
                    employee.location.location_name if employee.location else '', 
                    skills, 
                ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=index, column=col_num)
                    cell.value = cell_value

            file_path = os.path.join(settings.MEDIA_ROOT, 'employees.xlsx')
            workbook.save(file_path)

            # Assuming MEDIA_URL is configured properly in settings
            download_url = os.path.join(settings.MEDIA_URL, 'employees.xlsx')

            return JsonResponse({'download_url': download_url})

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

